from __future__ import annotations

import pytest
from actweave_knowledge.extraction.contracts import ExtractionError, HeaderRule


@pytest.mark.parametrize("row", [2, 5])
def test_tabular_explicit_rejects_nonexistent_or_continuation_physical_row(row):
    from actweave_knowledge.extraction.tabular import rows_to_documents

    with pytest.raises(ExtractionError, match="表头") as caught:
        rows_to_documents([(1, 3, ["a\nb\nc", "d"]), (4, 4, ["1", "2"])], sheet=None, rule=HeaderRule(mode="explicit", row=row))
    assert caught.value.reason_code == "HEADER_ROW_INVALID"


def test_tabular_auto_never_scans_a_record_beyond_tenth_physical_line():
    from actweave_knowledge.extraction.tabular import rows_to_documents

    docs = rows_to_documents([(1, 8, ["note"]), (9, 11, ["long\nheader\nname", "other"]), (12, 12, [1, 2])], sheet=None, rule=HeaderRule())
    assert all(d.kind == "table_row" for d in docs)
    assert not any(d.warnings for d in docs)


def test_csv_preserves_strings_multiline_commas_and_physical_locations(tmp_path):
    from actweave_knowledge.extraction.processor import ExtractProcessor
    from parsing_test_helpers import make_context, make_parse_profile, make_setting

    path = tmp_path / "values.csv"
    path.write_text('编号,标记,说明\n00123,NA,"上行,有逗号\n下行"\n00004,,正常\n', encoding="utf-8")
    profile = make_parse_profile(".csv", header_rules=(HeaderRule(mode="explicit", row=1),))
    docs = ExtractProcessor().extract(make_setting(path, profile=profile), make_context(tmp_path / "work"))
    data = [d for d in docs if d.kind == "table_row"]
    assert len(data) == 2
    assert data[0].page_content == "- 编号: 00123\n- 标记: NA\n- 说明: 上行,有逗号\n下行"
    assert data[0].source_spans[-1].location == {"row": 2, "row_end": 3, "column": 3, "encoding": "utf-8"}
    assert data[1].source_spans[-1].location["row"] == 4
    assert "None" not in data[1].page_content and "nan" not in data[1].page_content


@pytest.mark.parametrize("text", ["a,b\n1,2,3\n", "a,b\n1\n", 'a,b\n"unclosed,2\n', 'a,b\n"one"tail,2\n'])
def test_csv_bad_rows_fail_instead_of_disappearing(tmp_path, text):
    from actweave_knowledge.extraction.processor import ExtractProcessor
    from parsing_test_helpers import make_context, make_setting

    path = tmp_path / "broken.csv"
    path.write_text(text, encoding="utf-8")
    with pytest.raises(ExtractionError) as caught:
        ExtractProcessor().extract(make_setting(path), make_context(tmp_path / "work"))
    assert caught.value.reason_code == "CSV_ROW_INVALID"


def test_csv_notes_before_explicit_physical_header_can_have_different_width(tmp_path):
    from actweave_knowledge.extraction.processor import ExtractProcessor
    from parsing_test_helpers import make_context, make_parse_profile, make_setting

    path = tmp_path / "notes.csv"
    path.write_bytes('"介绍\r\n第二行"\r\na,b\r\n001,NA\r\n'.encode("utf-16"))
    profile = make_parse_profile(".csv", header_rules=(HeaderRule(mode="explicit", row=3),))
    docs = ExtractProcessor().extract(make_setting(path, profile=profile), make_context(tmp_path / "work"))
    assert docs[0].page_content == "介绍\n第二行"
    assert docs[-1].page_content == "- a: 001\n- b: NA"
    assert docs[-1].source_spans[-1].location["row"] == 4
    assert docs[-1].source_spans[-1].location["encoding"] == "utf-16"


def test_csv_empty_fields_and_leading_empty_lines_keep_source_rows(tmp_path):
    from actweave_knowledge.extraction.processor import ExtractProcessor
    from parsing_test_helpers import make_context, make_parse_profile, make_setting

    path = tmp_path / "none.csv"
    path.write_text("\n\n001,,NA\n002, null ,\n", encoding="utf-8")
    profile = make_parse_profile(".csv", header_rules=(HeaderRule(mode="none"),))
    docs = ExtractProcessor().extract(make_setting(path, profile=profile), make_context(tmp_path / "work"))
    assert len(docs) == 2 and docs[0].source_spans[-1].location["row"] == 3
    assert docs[0].page_content == "- 列 A: 001\n- 列 B: \n- 列 C: NA"
    assert docs[1].page_content == "- 列 A: 002\n- 列 B:  null \n- 列 C: "


def test_excel_blank_header_preserves_data_original_rows_and_header_cells(tmp_path):
    from actweave_knowledge.extraction.processor import ExtractProcessor
    from openpyxl import Workbook
    from parsing_test_helpers import make_context, make_setting

    path = tmp_path / "values.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "设备"
    for row in [["设备清单"], [], ["编号", None, "编号"], ["00123", "不能丢", "B"], [], ["00004", "NA", "C"]]:
        ws.append(row)
    wb.save(path)
    wb.close()
    docs = ExtractProcessor().extract(make_setting(path), make_context(tmp_path / "work"))
    data = [d for d in docs if d.kind == "table_row"]
    assert len(data) == 2
    assert data[0].page_content == "- 编号 [A]: 00123\n- 列 B: 不能丢\n- 编号 [C]: B"
    assert [d.source_spans[-1].location["row"] for d in data] == [4, 6]
    header = next(d for d in docs if d.kind == "table_header")
    assert [header.page_content[s.start : s.end] for s in header.source_spans] == ["编号", "", "编号"]
    assert any("设备清单" in d.page_content for d in docs)


def test_excel_sheet_rules_numeric_tables_and_empty_leading_rows(tmp_path):
    from actweave_knowledge.extraction.processor import ExtractProcessor
    from openpyxl import Workbook
    from parsing_test_helpers import make_context, make_parse_profile, make_setting

    path = tmp_path / "sheets.xlsx"
    wb = Workbook()
    wb.active.title = "数字"
    for row in [[], [1, 2], [3, 4]]:
        wb.active.append(row)
    wb.create_sheet("无表头").append(["first", "value"])
    explicit = wb.create_sheet("指定")
    for row in [[], ["note"], ["ID", "Value"], ["001", "NA"]]:
        explicit.append(row)
    wb.save(path)
    wb.close()
    profile = make_parse_profile(".xlsx", header_rules=(HeaderRule(sheet="无表头", mode="none"), HeaderRule(sheet="指定", mode="explicit", row=3)))
    docs = ExtractProcessor().extract(make_setting(path, profile=profile), make_context(tmp_path / "work"))
    data = [d for d in docs if d.kind == "table_row"]
    assert [(d.source_spans[-1].location["sheet"], d.source_spans[-1].location["row"]) for d in data] == [("数字", 2), ("数字", 3), ("无表头", 1), ("指定", 4)]
    assert data[0].page_content == "- 列 A: 1\n- 列 B: 2"
    assert data[-1].page_content == "- ID: 001\n- Value: NA"


def test_excel_missing_formula_cache_warns_and_never_evaluates(tmp_path):
    from actweave_knowledge.extraction.processor import ExtractProcessor
    from openpyxl import Workbook
    from parsing_test_helpers import make_context, make_setting

    path = tmp_path / "formula.xlsx"
    wb = Workbook()
    wb.active.append(["ID", "Amount"])
    wb.active.append(["001", "=1+2"])
    wb.active.append([None, "=4+5"])
    wb.save(path)
    wb.close()
    docs = ExtractProcessor().extract(make_setting(path), make_context(tmp_path / "work"))
    data = [d for d in docs if d.kind == "table_row"]
    assert data[0].page_content == "- ID: 001\n- Amount: "
    warnings = [w for d in docs for w in d.warnings if w.code == "FORMULA_CACHE_MISSING"]
    assert [(w.source_position["row"], w.source_position["column"]) for w in warnings] == [(2, 2), (3, 2)]
    assert not any("=1+2" in d.page_content or "=4+5" in d.page_content for d in docs)


def test_excel_images_keep_all_sheets_and_anchors_including_notes_and_blank_columns(tmp_path):
    import io

    from actweave_knowledge.extraction.processor import ExtractProcessor
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as SheetImage
    from parsing_test_helpers import make_context, make_setting
    from PIL import Image

    png = io.BytesIO()
    Image.new("RGB", (2, 3), "red").save(png, format="PNG")
    path = tmp_path / "images.xlsx"
    wb = Workbook()
    wb.active.title = "设备"
    wb.active.append(["介绍"])
    wb.active.append(["编号", None, "说明"])
    wb.active.append(["001", None, "内容"])
    for anchor in ["B1", "B3", "E7"]:
        wb.active.add_image(SheetImage(io.BytesIO(png.getvalue())), anchor)
    wb.create_sheet("仅图片").add_image(SheetImage(io.BytesIO(png.getvalue())), "C4")
    wb.save(path)
    wb.close()
    context = make_context(tmp_path / "work")
    docs = ExtractProcessor().extract(make_setting(path), context)
    occurrences = [a for d in docs for a in d.attachments]
    assert [(a.source.location["sheet"], a.source.location["row"], a.source.location["column"]) for a in occurrences] == [("设备", 1, 2), ("设备", 3, 2), ("设备", 7, 5), ("仅图片", 4, 3)]
    assert len(context.sink.assets) == 1 and len(context.sink.occurrences) == 4
    assert len({a.ref for a in occurrences}) == 1
    for doc in docs:
        for occurrence in doc.attachments:
            assert doc.page_content[occurrence.source.start : occurrence.source.end] == f"![{occurrence.alt_text}](knowledge-attachment:{occurrence.ref})"


def test_xls_real_empty_rows_do_not_renumber_source_cells(tmp_path):
    from pathlib import Path

    from actweave_knowledge.extraction.processor import ExtractProcessor
    from parsing_test_helpers import make_context, make_parse_profile, make_setting

    # Unmodified xlrd 2.0.1 fixture, commit b8d573e11ec149da695d695c81a156232b89a949.
    path = Path(__file__).parent / "fixtures" / "xlrd" / "issue20.xls"
    profile = make_parse_profile(".xls", header_rules=(HeaderRule(mode="none"),))
    docs = ExtractProcessor().extract(make_setting(path, profile=profile), make_context(tmp_path / "work"))
    data = [d for d in docs if d.kind == "table_row"]
    assert [d.source_spans[-1].location["row"] for d in data] == [1, 2, 3, 4, 5, 11, 12]
    assert data[-2].source_spans[-1].location == {"sheet": "Sheet1", "row": 11, "column": 12}
    assert data[-2].page_content.endswith("- 列 L: asasa")


def test_csv_field_above_stdlib_default_limit_is_not_rejected(tmp_path):
    from actweave_knowledge.extraction.processor import ExtractProcessor
    from parsing_test_helpers import make_context, make_parse_profile, make_setting

    path = tmp_path / "large.csv"
    value = "x" * 150_000
    path.write_text("id,value\n001," + value + "\n", encoding="utf-8")
    profile = make_parse_profile(".csv", header_rules=(HeaderRule(mode="explicit", row=1),))
    docs = ExtractProcessor().extract(make_setting(path, profile=profile), make_context(tmp_path / "work"))
    source = docs[-1].source_spans[-1]
    assert docs[-1].page_content[source.start : source.end] == value


def test_excel_cached_formula_value_is_kept_without_missing_cache_warning(tmp_path):
    import zipfile

    from actweave_knowledge.extraction.processor import ExtractProcessor
    from openpyxl import Workbook
    from parsing_test_helpers import make_context, make_setting

    original = tmp_path / "original.xlsx"
    path = tmp_path / "cached.xlsx"
    wb = Workbook()
    wb.active.append(["ID", "Amount"])
    wb.active.append(["001", "=1+2"])
    wb.save(original)
    wb.close()
    with zipfile.ZipFile(original) as archive, zipfile.ZipFile(path, "w") as target:
        for item in archive.infolist():
            payload = archive.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                payload = payload.replace(b"<f>1+2</f><v></v>", b"<f>1+2</f><v>3</v>")
            target.writestr(item, payload)
    docs = ExtractProcessor().extract(make_setting(path), make_context(tmp_path / "work"))
    assert docs[-1].page_content == "- ID: 001\n- Amount: 3"
    assert not any(w.code == "FORMULA_CACHE_MISSING" for d in docs for w in d.warnings)


def test_excel_cell_hyperlinks_keep_safe_target_and_visible_value(tmp_path):
    from actweave_knowledge.extraction.processor import ExtractProcessor
    from openpyxl import Workbook
    from parsing_test_helpers import make_context, make_setting

    path = tmp_path / "links.xlsx"
    wb = Workbook()
    wb.active.append(["ID", "Docs"])
    wb.active.append(["001", "Documentation"])
    wb.active["B2"].hyperlink = "https://example.com/docs"
    wb.active.append(["002", "unsafe label"])
    wb.active["B3"].hyperlink = "javascript:alert(1)"
    wb.save(path)
    wb.close()
    docs = ExtractProcessor().extract(make_setting(path), make_context(tmp_path / "work"))
    assert "[Documentation](https://example.com/docs)" in docs[1].page_content
    assert "unsafe label" in docs[2].page_content and "javascript:" not in docs[2].page_content


def test_excel_corrupt_media_is_visible_and_structured_without_raw_warning(tmp_path, monkeypatch):
    import io
    import warnings
    import zipfile

    from actweave_knowledge.extraction.builtin import excel_extractor
    from actweave_knowledge.extraction.processor import ExtractProcessor
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as SheetImage
    from parsing_test_helpers import make_context, make_setting
    from PIL import Image

    original = tmp_path / "original.xlsx"
    path = tmp_path / "broken-image.xlsx"
    image_data = io.BytesIO()
    Image.new("RGB", (2, 3), "red").save(image_data, format="PNG")
    workbook = Workbook()
    workbook.active.title = "设备"
    workbook.active.append(["id", "value"])
    workbook.active.append(["001", "content survives"])
    workbook.active.add_image(SheetImage(io.BytesIO(image_data.getvalue())), "C2")
    workbook.save(original)
    workbook.close()
    with zipfile.ZipFile(original) as archive, zipfile.ZipFile(path, "w") as target:
        for item in archive.infolist():
            payload = b"not-a-png-image" if item.filename == "xl/media/image1.png" else archive.read(item)
            target.writestr(item, payload)

    loaded, closed = [], []
    real_load = excel_extractor.load_workbook

    def tracking_load(*args, **kwargs):
        wb = real_load(*args, **kwargs)
        loaded.append(wb)
        real_close = wb.close

        def tracking_close():
            real_close()
            closed.append(wb)

        wb.close = tracking_close
        return wb

    monkeypatch.setattr(excel_extractor, "load_workbook", tracking_load)
    with warnings.catch_warnings(record=True) as native_warnings:
        warnings.simplefilter("always")
        docs = ExtractProcessor().extract(make_setting(path), make_context(tmp_path / "work"))
    assert len(loaded) == 2 and len(closed) == 2 and {id(wb) for wb in closed} == {id(wb) for wb in loaded}
    assert any("content survives" in d.page_content for d in docs)
    bad = [d for d in docs if any(w.code == "IMAGE_CORRUPT" for w in d.warnings)]
    assert len(bad) == 1
    assert bad[0].page_content == "（图片无法安全解码）"
    assert not bad[0].attachments
    assert bad[0].warnings[0].source_position == {"sheet": "设备", "row": 2, "column": 3, "image_index": 1}
    assert bad[0].source_spans[0].role == "context_prefix"
    assert bad[0].page_content[bad[0].source_spans[0].start : bad[0].source_spans[0].end] == bad[0].page_content
    assert not native_warnings
    assert "xl/media" not in str([d.model_dump() for d in docs])


def test_excel_sink_image_rejection_keeps_placeholder_and_cleans_tempfile(tmp_path):
    import io

    from actweave_knowledge.extraction.contracts import ExtractionLimits
    from actweave_knowledge.extraction.images import LocalAttachmentSink
    from actweave_knowledge.extraction.processor import ExtractProcessor
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as SheetImage
    from parsing_test_helpers import make_context, make_setting
    from PIL import Image

    path = tmp_path / "large-image.xlsx"
    image_data = io.BytesIO()
    Image.new("RGB", (2, 3), "red").save(image_data, format="PNG")
    workbook = Workbook()
    workbook.active.append(["id", "value"])
    workbook.active.append(["001", "survives"])
    workbook.active.add_image(SheetImage(io.BytesIO(image_data.getvalue())), "B2")
    workbook.save(path)
    workbook.close()
    context = make_context(tmp_path / "work")
    sink = LocalAttachmentSink(context.work_dir, ExtractionLimits(max_image_pixels=1))
    docs = ExtractProcessor().extract(make_setting(path), context.model_copy(update={"sink": sink}))
    bad = next(d for d in docs if any(w.code == "IMAGE_LIMIT_EXCEEDED" for w in d.warnings))
    assert bad.page_content == "（图片超过安全上限）"
    assert bad.source_spans[0].role == "context_prefix"
    assert bad.warnings[0].source_position["row"] == 2
    assert not bad.attachments and not sink.assets
    assert not list(context.work_dir.glob("excel-image-*"))


def test_excel_sink_io_failure_remains_fatal_and_closes_workbooks(tmp_path, monkeypatch):
    import io

    from actweave_knowledge.extraction.builtin import excel_extractor
    from actweave_knowledge.extraction.processor import ExtractProcessor
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as SheetImage
    from parsing_test_helpers import CollectingAttachmentSink, make_context, make_setting
    from PIL import Image

    path = tmp_path / "image-io.xlsx"
    image_data = io.BytesIO()
    Image.new("RGB", (2, 3), "red").save(image_data, format="PNG")
    workbook = Workbook()
    workbook.active.add_image(SheetImage(io.BytesIO(image_data.getvalue())), "B2")
    workbook.save(path)
    workbook.close()
    loaded, closed = [], []
    real_load = excel_extractor.load_workbook

    def tracking_load(*args, **kwargs):
        wb = real_load(*args, **kwargs)
        loaded.append(wb)
        real_close = wb.close

        def tracking_close():
            real_close()
            closed.append(wb)

        wb.close = tracking_close
        return wb

    def permission_denied(*args, **kwargs):
        raise PermissionError(13, "permission denied")

    monkeypatch.setattr(excel_extractor, "load_workbook", tracking_load)
    monkeypatch.setattr(CollectingAttachmentSink, "accept", permission_denied)
    context = make_context(tmp_path / "work")
    with pytest.raises(PermissionError):
        ExtractProcessor().extract(make_setting(path), context)
    assert len(loaded) == 2 and len(closed) == 2 and {id(wb) for wb in closed} == {id(wb) for wb in loaded}
    assert not list(context.work_dir.glob("excel-image-*"))


def test_xlsx_image_survives_normalization_and_manifest_closure(tmp_path):
    import hashlib

    from actweave_knowledge.extraction.contracts import ExtractionResult
    from actweave_knowledge.extraction.images import LocalAttachmentSink
    from actweave_knowledge.extraction.manifest import canonical_parse_fingerprint, decode_manifest, encode_manifest
    from actweave_knowledge.extraction.normalizer import normalize_documents
    from actweave_knowledge.extraction.processor import ExtractProcessor
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as SheetImage
    from parsing_test_helpers import make_context, make_setting
    from PIL import Image

    png = tmp_path / "red.png"
    Image.new("RGB", (3, 2), "red").save(png)
    workbook = Workbook()
    workbook.active.append(["Name", "Value"])
    workbook.active.append(["row", "00123"])
    workbook.active.add_image(SheetImage(png), "C2")
    path = tmp_path / "source.xlsx"
    workbook.save(path)
    setting = make_setting(path)
    context = make_context(tmp_path / "work")
    sink = LocalAttachmentSink(context.work_dir, context.limits)
    documents = normalize_documents(ExtractProcessor().extract(setting, context.model_copy(update={"sink": sink})))
    result = ExtractionResult(
        documents=tuple(documents), attachments=tuple(asset.attachment for asset in sink.assets), source_sha256=hashlib.sha256(path.read_bytes()).hexdigest(), parse_fingerprint=canonical_parse_fingerprint(setting.profile)
    )
    assert len(result.attachments) == 1
    decoded = decode_manifest(encode_manifest(result), context.limits)
    occurrences = [occurrence for doc in decoded.documents for occurrence in doc.attachments]
    assert len(occurrences) == 1
    assert occurrences[0].ref == result.attachments[0].ref
    assert occurrences[0].source.location == {"sheet": "Sheet", "image_index": 1, "row": 2, "column": 3}
    assert not any(w.code == "EXTERNAL_IMAGE_NOT_FETCHED" for doc in decoded.documents for w in doc.warnings)
