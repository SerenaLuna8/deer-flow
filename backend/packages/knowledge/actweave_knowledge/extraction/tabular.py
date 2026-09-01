"""Shared table headers and field binding without discarding source cells."""

from __future__ import annotations

from collections import Counter

from openpyxl.utils.cell import get_column_letter

from .contracts import Document, ExtractionError, HeaderRule, ParseWarning, SourceSpan

TableRow = tuple[int, int, list[object]]


def select_header(rows: list[list[object]], rule: HeaderRule) -> int | None:
    """Select a one-based row in a physical-row view, never guess numeric headers."""
    if rule.mode == "none":
        return None
    if rule.mode == "explicit":
        if rule.row is None or rule.row > len(rows):
            raise ExtractionError("HEADER_ROW_INVALID", "表头行不存在")
        return rule.row
    for number, row in enumerate(rows[:10], 1):
        if sum(isinstance(value, str) and bool(value.strip()) for value in row) >= 2:
            return number
    return None


def header_record_index(rows: list[TableRow], rule: HeaderRule) -> int | None:
    """Map physical header rows to zero-based records (quoted CSV can span rows)."""
    if rule.mode == "explicit":
        for index, (start, _end, _values) in enumerate(rows):
            if start == rule.row:
                return index
        raise ExtractionError("HEADER_ROW_INVALID", "表头行不存在或不是记录起始行")
    if rule.mode == "none":
        return None
    physical_rows: list[list[object]] = [[] for _ in range(10)]
    for start, end, values in rows:
        if start > 10:
            break
        if end <= 10:
            physical_rows[start - 1] = values
    number = select_header(physical_rows, rule)
    return next((index for index, row in enumerate(rows) if row[0] == number), None)


def column_labels(values: list[object]) -> list[str]:
    """Fill blank names and disambiguate duplicate names by Excel column letter."""
    names = [str(value).strip() if value is not None and str(value).strip() else f"列 {get_column_letter(index)}" for index, value in enumerate(values, 1)]
    counts = Counter(names)
    return [f"{name} [{get_column_letter(index)}]" if counts[name] > 1 else name for index, name in enumerate(names, 1)]


def _text(value: object) -> str:
    return "" if value is None else str(value)


def _location(row: TableRow, sheet: str | None, column: int) -> dict[str, str | int]:
    location: dict[str, str | int] = {"row": row[0], "column": column}
    if sheet is None:
        location["row_end"] = row[1]
    else:
        location["sheet"] = sheet
    return location


def _block_id(row: TableRow, sheet: str | None) -> str:
    return f"table:{sheet or ''}:row:{row[0]}"


def _raw_document(row: TableRow, sheet: str | None, kind: str, warnings: tuple[ParseWarning, ...] = ()) -> Document:
    parts = [_text(value) for value in row[2]]
    spans = []
    offset = 0
    for column, part in enumerate(parts, 1):
        spans.append(SourceSpan(block_id=_block_id(row, sheet), start=offset, end=offset + len(part), location=_location(row, sheet, column)))
        offset += len(part) + 1
    return Document(page_content="\t".join(parts), source_spans=tuple(spans), kind=kind, warnings=warnings)


def rows_to_documents(rows: list[TableRow], *, sheet: str | None, rule: HeaderRule) -> list[Document]:
    """Keep context/header cells and render each data row with field/value spans."""
    header_index = header_record_index(rows, rule)
    header = rows[header_index] if header_index is not None else None
    table_rows = rows[header_index:] if header_index is not None else rows
    width = max((len(row[2]) for row in table_rows), default=0)
    labels = column_labels((header[2] if header is not None else []) + [None] * (width - (len(header[2]) if header is not None else 0)))
    documents = []
    for index, row in enumerate(rows):
        if index == header_index:
            warnings = ()
            if rule.mode == "auto":
                warnings = (ParseWarning(code="HEADER_INFERRED", message="已自动识别表头，请确认", source_position=_location(row, sheet, 1)),)
            documents.append(_raw_document(row, sheet, "table_header", warnings))
            continue
        if not any(value is not None and _text(value) != "" for value in row[2]):
            continue
        if header_index is not None and index < header_index:
            documents.append(_raw_document(row, sheet, "context"))
            continue
        parts, spans = [], []
        offset = 0
        for column, label in enumerate(labels, 1):
            value = _text(row[2][column - 1]) if column <= len(row[2]) else ""
            prefix = f"- {label}: "
            header_source = header if header is not None and column <= len(header[2]) else row
            spans.append(SourceSpan(block_id=_block_id(header_source, sheet), start=offset, end=offset + len(prefix), location=_location(header_source, sheet, column), role="context_prefix"))
            spans.append(SourceSpan(block_id=_block_id(row, sheet), start=offset + len(prefix), end=offset + len(prefix) + len(value), location=_location(row, sheet, column)))
            parts.append(prefix + value)
            offset += len(prefix) + len(value) + 1
        documents.append(Document(page_content="\n".join(parts), source_spans=tuple(spans), kind="table_row"))
    return documents
