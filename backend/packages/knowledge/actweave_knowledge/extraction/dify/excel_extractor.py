"""Excel document extractor adapted from Dify's pinned ExcelExtractor.

Upstream: api/core/rag/extractor/excel_extractor.py at
9c16c865977e9d89a9ec7ae0536e893f4385a758. See UPSTREAM.md and patches.md.
Workbook/sheet/cell traversal and drawing byte/anchor extraction are retained;
header selection, source positions, and injected asset delivery replace the
upstream lossy column filtering and host database/storage coupling.
"""

from __future__ import annotations

import re
import tempfile
import warnings
from pathlib import Path
from typing import override
from urllib.parse import quote, urlsplit
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
from openpyxl.packaging.relationship import get_dependents, get_rels_path
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.xml.functions import fromstring

from ..base import BaseExtractor
from ..contracts import AttachmentOccurrence, Document, ExtractionContext, ExtractionError, ExtractSetting, HeaderRule, ParseWarning, SourceSpan
from ..images import ImageRejected
from ..tabular import TableRow, header_record_index, rows_to_documents


class ExcelExtractor(BaseExtractor):
    """Load Excel files while preserving sheet names and original cell locations."""

    @override
    def extract(self, setting: ExtractSetting, context: ExtractionContext) -> list[Document]:
        """Load from Excel file in xls or xlsx format using xlrd and openpyxl."""
        context.check_cancelled()
        documents: list[Document] = []
        file_extension = Path(setting.original_name).suffix.lower()
        if file_extension == ".xlsx":
            # Worksheet drawing objects, including embedded images, are not available in read-only mode.
            with warnings.catch_warnings(record=True) as load_warnings:
                warnings.simplefilter("always")
                wb = load_workbook(setting.source_path, data_only=True, read_only=False)
            formulas = None
            try:
                discarded_images = self._discarded_image_documents(setting.source_path, wb, load_warnings)
                formulas = load_workbook(setting.source_path, data_only=False, read_only=True)
                for sheet_name in wb.sheetnames:
                    context.check_cancelled()
                    sheet = wb[sheet_name]
                    rows: list[TableRow] = []
                    for row_number, row in enumerate(sheet.iter_rows(values_only=False), 1):
                        context.check_cancelled()
                        rows.append((row_number, row_number, [cell.value for cell in row]))
                    rule = self._header_rule(setting, sheet_name)
                    header_index = header_record_index(rows, rule)
                    # Infer from native cells before adding hyperlink Markdown; raw headers stay intact.
                    for index, (_start, _end, values) in enumerate(rows):
                        if header_index is not None and index <= header_index:
                            continue
                        for column, value in enumerate(values, 1):
                            cell = sheet.cell(index + 1, column)
                            if hasattr(cell, "hyperlink") and cell.hyperlink:
                                target = getattr(cell.hyperlink, "target", None)
                                if target and urlsplit(target).scheme.lower() in {"http", "https", "mailto"}:
                                    display_value = str(value) if value is not None and str(value).strip() else target
                                    display_value = display_value.replace("\\", "\\\\").replace("[", "\\[").replace("]", "\\]")
                                    values[column - 1] = f"[{display_value}]({quote(target, safe=":/?#@!$&'*+,;=%")})"
                    selected_rule = rule if header_index is not None else HeaderRule(sheet=sheet_name, mode="none")
                    sheet_docs = rows_to_documents(rows, sheet=sheet_name, rule=selected_rule)
                    sheet_docs = self._formula_warnings(sheet_docs, sheet, formulas[sheet_name])
                    documents.extend(sheet_docs)
                    documents.extend(self._extract_images_from_sheet(sheet_name=sheet_name, sheet=sheet, context=context))
                    documents.extend(discarded_images.get(sheet_name, []))
                documents.extend(discarded_images.get(None, []))
            finally:
                if formulas is not None:
                    formulas.close()
                wb.close()
        elif file_extension == ".xls":
            import xlrd

            excel_file = xlrd.open_workbook(setting.source_path, on_demand=True)
            try:
                for excel_sheet_name in excel_file.sheet_names():
                    context.check_cancelled()
                    sheet = excel_file.sheet_by_name(excel_sheet_name)
                    rows = [(number + 1, number + 1, list(sheet.row_values(number))) for number in range(sheet.nrows)]
                    documents.extend(rows_to_documents(rows, sheet=excel_sheet_name, rule=self._header_rule(setting, excel_sheet_name)))
            finally:
                excel_file.release_resources()
        else:
            raise ExtractionError("FORMAT_UNSUPPORTED", "不支持的表格格式")
        return documents

    @staticmethod
    def _header_rule(setting: ExtractSetting, sheet_name: str) -> HeaderRule:
        rules = setting.profile.header_rules
        return next((rule for rule in rules if rule.sheet == sheet_name), next((rule for rule in rules if rule.sheet is None), HeaderRule(sheet=sheet_name)))

    @staticmethod
    def _formula_warnings(documents: list[Document], sheet: Worksheet, formulas) -> list[Document]:
        missing: dict[int, list[ParseWarning]] = {}
        for row in formulas.iter_rows():
            for cell in row:
                if cell.data_type == "f" and sheet.cell(cell.row, cell.column).value is None:
                    warning = ParseWarning(code="FORMULA_CACHE_MISSING", message="公式缺少已保存的计算结果，保留空值", source_position={"sheet": sheet.title, "row": cell.row, "column": cell.column})
                    missing.setdefault(cell.row, []).append(warning)
        result = []
        for document in documents:
            row_number = document.source_spans[0].location["row"]
            # Data documents begin with a header prefix; locate their real source row.
            source = next((span for span in document.source_spans if span.role == "source"), None)
            if source is not None:
                row_number = source.location["row"]
            warnings = tuple(missing.pop(row_number, []))
            result.append(document.model_copy(update={"warnings": document.warnings + warnings}))
        # A formula-only row with no cached values still needs a visible warning/source.
        for row_number, warnings in missing.items():
            source = SourceSpan(block_id=f"table:{sheet.title}:row:{row_number}", start=0, end=0, location=warnings[0].source_position)
            result.append(Document(page_content="", source_spans=(source,), kind="table_row", warnings=tuple(warnings)))
        result.sort(key=lambda document: next(span.location["row"] for span in document.source_spans if span.role == "source"))
        return result

    @staticmethod
    def _image_placeholder(source: SourceSpan, warning: ParseWarning) -> Document:
        text = f"（{warning.message}）"
        source = source.model_copy(update={"start": 0, "end": len(text), "role": "context_prefix"})
        warning = warning.model_copy(update={"source_position": dict(source.location)})
        return Document(page_content=text, source_spans=(source,), kind="image", warnings=(warning,))

    def _discarded_image_documents(self, source_path: Path, workbook, observed: list[warnings.WarningMessage]) -> dict[str | None, list[Document]]:
        """Map openpyxl's image-loss warnings to drawing metadata, without decoding media."""
        removed_targets: set[str] = set()
        unknown_losses = 0
        for item in observed:
            removed = re.fullmatch(r"The image (.+) will be removed because it cannot be read", str(item.message))
            if removed:
                removed_targets.add(removed.group(1))
            elif re.fullmatch(r".+ image format is not supported so the image is being dropped", str(item.message)):
                unknown_losses += 1
            else:
                # Do not mistake unrelated workbook warnings for a corrupt image.
                warnings.warn_explicit(item.message, item.category, item.filename, item.lineno)
        documents: dict[str | None, list[Document]] = {}
        if not removed_targets and not unknown_losses:
            return documents
        matched = set()
        with ZipFile(source_path) as archive:
            for sheet in workbook.worksheets:
                image_index = 0
                for rel in sheet._rels.find(SpreadsheetDrawing._rel_type):
                    drawing = SpreadsheetDrawing.from_tree(fromstring(archive.read(rel.target)))
                    dependencies = get_dependents(archive, get_rels_path(rel.target))
                    for image_rel in drawing._blip_rels:
                        image_index += 1
                        target = dependencies.get(image_rel.embed).target
                        if target not in removed_targets:
                            continue
                        matched.add(target)
                        location: dict[str, str | int] = {"sheet": sheet.title, "image_index": image_index}
                        marker = getattr(image_rel.anchor, "_from", None)
                        if marker is not None:
                            location.update(row=marker.row + 1, column=marker.col + 1)
                        source = SourceSpan(block_id=f"table:{sheet.title}:discarded-image:{image_index}", start=0, end=0, location=location)
                        warning = ParseWarning(code="IMAGE_CORRUPT", message="图片无法安全解码")
                        documents.setdefault(sheet.title, []).append(self._image_placeholder(source, warning))
        # A library loss with no recoverable anchor remains visible at document scope.
        for number in range(len(removed_targets - matched) + unknown_losses):
            source = SourceSpan(block_id=f"table:unlocated-image:{number + 1}", start=0, end=0)
            warning = ParseWarning(code="IMAGE_CORRUPT", message="图片无法安全解码")
            documents.setdefault(None, []).append(self._image_placeholder(source, warning))
        return documents

    def _extract_images_from_sheet(self, sheet_name: str, sheet: Worksheet, context: ExtractionContext) -> list[Document]:
        """Extract every worksheet image, retaining each original anchor occurrence."""
        images = getattr(sheet, "_images", None) or []
        documents = []
        for image_index, image in enumerate(images, 1):
            context.check_cancelled()
            marker = getattr(getattr(image, "anchor", None), "_from", None)
            row_idx = getattr(marker, "row", None)
            col_idx = getattr(marker, "col", None)
            location: dict[str, str | int] = {"sheet": sheet_name, "image_index": image_index}
            warnings = ()
            if row_idx is not None and col_idx is not None:
                location.update(row=row_idx + 1, column=col_idx + 1)
            else:
                warnings = (ParseWarning(code="IMAGE_ANCHOR_UNAVAILABLE", message="图片未提供单元格锚点", source_position=location),)
            source = SourceSpan(block_id=f"table:{sheet_name}:image:{image_index}", start=0, end=0, location=location)
            alt_text = f"工作表图片 {image_index}"
            try:
                image_bytes = self._get_image_bytes(image)
                with tempfile.NamedTemporaryFile(dir=context.work_dir, prefix="excel-image-", suffix=".bin", delete=False) as output:
                    path = Path(output.name)
                    output.write(image_bytes)
                try:
                    attachment = context.sink.accept(path, alt_text=alt_text, source=source)
                finally:
                    path.unlink(missing_ok=True)
            except ImageRejected as error:
                documents.append(self._image_placeholder(source, error.warning))
                continue
            content = f"![{alt_text}](knowledge-attachment:{attachment.ref})"
            source = source.model_copy(update={"end": len(content)})
            occurrence = AttachmentOccurrence(ref=attachment.ref, alt_text=alt_text, source=source)
            documents.append(Document(page_content=content, source_spans=(source,), kind="image", attachments=(occurrence,), warnings=warnings))
        return documents

    def _get_image_bytes(self, image) -> bytes:
        """Return embedded image bytes from an openpyxl image object."""
        data_loader = getattr(image, "_data", None)
        if not callable(data_loader):
            raise ImageRejected(ParseWarning(code="IMAGE_CORRUPT", message="图片无法安全解码"))
        try:
            data = data_loader()
            if isinstance(data, bytes):
                return data
            if isinstance(data, bytearray):
                return bytes(data)
        except (OSError, ValueError, SyntaxError) as error:
            if isinstance(error, OSError) and error.errno is not None:
                raise
            raise ImageRejected(ParseWarning(code="IMAGE_CORRUPT", message="图片无法安全解码")) from None
        raise ImageRejected(ParseWarning(code="IMAGE_CORRUPT", message="图片无法安全解码"))
